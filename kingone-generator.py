import pandas as pd
import openpyxl
from openpyxl.styles import Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import itertools
import random
import sys
import os

class PadelTournamentGenerator:
    def __init__(self, duration_minutes, match_duration, start_time, num_courts, num_players, export_path):
        self.duration_minutes = duration_minutes
        self.match_duration = match_duration
        self.start_time = start_time
        self.num_courts = num_courts
        self.num_players = num_players
        self.export_path = export_path
        
        # Validate inputs
        self.validate_inputs()
        
        # Calculate tournament parameters
        self.num_rounds = self.duration_minutes // self.match_duration
        self.total_matches = self.num_rounds * self.num_courts
        
        # Generate players
        self.players = self.generate_players()
        
        # Generate matches
        self.matches = self.generate_matches()
    
    def validate_inputs(self):
        """Validate input parameters."""
        if self.num_players % 4 != 0:
            raise ValueError(f"Il numero di giocatori ({self.num_players}) deve essere multiplo di 4.")
        
        courts_needed = self.num_players // 4
        if courts_needed > self.num_courts:
            raise ValueError(f"Non ci sono abbastanza campi ({self.num_courts}) per il numero di giocatori ({self.num_players}). Servono almeno {courts_needed} campi.")
    
    def generate_players(self):
        """Generate a list of players, half right-handed (D) and half left-handed (S)."""
        half_players = self.num_players // 2
        right_handed = [f"D{i+1}" for i in range(half_players)]
        left_handed = [f"S{i+1}" for i in range(half_players)]
        
        return {"right": right_handed, "left": left_handed}
    
    def generate_matches(self):
        """Generate tournament matches ensuring all players play in each round."""
        matches = []
        right_players = self.players["right"]
        left_players = self.players["left"]
        
        # Track player pairing history
        player_pairings = {}
        for r in right_players:
            for l in left_players:
                player_pairings[(r, l)] = 0  # Initialize count of times these players are paired
        
        # Track player opposition history
        player_oppositions = {}
        for r1 in right_players:
            for r2 in right_players:
                if r1 != r2:
                    player_oppositions[(r1, r2)] = 0
        for l1 in left_players:
            for l2 in left_players:
                if l1 != l2:
                    player_oppositions[(l1, l2)] = 0
        
        # Generate matches for each round
        current_time = self.start_time
        for round_num in range(1, self.num_rounds + 1):
            # For each round, create a copy of players to track who hasn't played yet
            available_right = right_players.copy()
            available_left = left_players.copy()
            
            # Number of courts to use for this round
            courts_needed = min(self.num_courts, self.num_players // 4)
            
            # Create matches for this round
            for court in range(1, courts_needed + 1):
                # Get optimal players for this match from available players
                match_players = self.select_optimal_players_for_round(
                    available_right, available_left, player_pairings, player_oppositions
                )
                
                if not match_players:
                    break
                
                r1, r2, l1, l2 = match_players
                
                # Update pairings count
                player_pairings[(r1, l1)] += 1
                player_pairings[(r2, l2)] += 1
                
                # Update oppositions count
                player_oppositions[(r1, r2)] += 1
                player_oppositions[(l1, l2)] += 1
                
                # Create match entry
                match = {
                    "round": round_num,
                    "time": current_time.strftime("%H:%M"),
                    "court": court,
                    "player1": r1,  # Right-handed
                    "player2": l1,  # Left-handed
                    "player3": r2,  # Right-handed
                    "player4": l2,  # Left-handed
                    "score_team1": 0,
                    "score_team2": 0
                }
                matches.append(match)
                
                # Remove used players from available lists
                available_right.remove(r1)
                available_right.remove(r2)
                available_left.remove(l1)
                available_left.remove(l2)
            
            # Update time for next round
            current_time += timedelta(minutes=self.match_duration)
        
        return matches
    
    def select_optimal_players_for_round(self, available_right, available_left, player_pairings, player_oppositions):
        """Select optimal players from available players for the current round."""
        if len(available_right) < 2 or len(available_left) < 2:
            return None
        
        # Find optimal pairings
        min_pairing_score = float('inf')
        best_combination = None
        
        # Try different combinations of right and left players
        for r1, r2 in itertools.combinations(available_right, 2):
            for l1, l2 in itertools.combinations(available_left, 2):
                # Calculate a score based on how often these players have been paired or opposed
                pairing_score = (
                    player_pairings[(r1, l1)] + 
                    player_pairings[(r2, l2)] +
                    player_oppositions[(r1, r2)] + 
                    player_oppositions[(l1, l2)]
                )
                
                if pairing_score < min_pairing_score:
                    min_pairing_score = pairing_score
                    best_combination = (r1, r2, l1, l2)
        
        return best_combination
    
    def generate_excel(self):
        """Generate Excel file with tournament details."""
        # Create Excel writer using openpyxl
        writer = pd.ExcelWriter(self.export_path, engine='openpyxl')
        
        # Create players dataframe
        players_data = []
        for i, player in enumerate(self.players["right"]):
            players_data.append({"NOME": player, "POSIZIONE": "DESTRO"})
        for i, player in enumerate(self.players["left"]):
            players_data.append({"NOME": player, "POSIZIONE": "SINISTRO"})
        
        players_df = pd.DataFrame(players_data)
        players_df.to_excel(writer, sheet_name="Giocatori", index=False)
        
        # Create matches dataframe
        matches_df = pd.DataFrame(self.matches)
        matches_df = matches_df[["round", "time", "court", "player1", "player2", "player3", "player4", "score_team1", "score_team2"]]
        matches_df.columns = ["TURNO", "ORARIO PARTITA", "CAMPO", "GIOCATORE 1(Destro)", "GIOCATORE 2(Sinistro)", 
                             "GIOCATORE 3(Destro)", "GIOCATORE 4(Sinistro)", "GAMES VINTO SQUADRA 1", "GAMES VINTO SQUADRA 2"]
        matches_df.to_excel(writer, sheet_name="Partite", index=False)
        
        # Calculate player summary
        all_players = self.players["right"] + self.players["left"]
        summary_data = []
        
        for player in all_players:
            summary_data.append({"NOME": player, "GAMES VINTI": 0, "PARTITE VINTI": 0,"GAMES PERSI": 0, "GAMES DIFFERENZA": 0})  # Will be replaced with formulas
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name="Reassunto", index=False)
        
        # Save and get workbook for additional formatting
        writer.close()
        
        # Now use openpyxl to add cell references and protection
        wb = openpyxl.load_workbook(self.export_path)
        
        # Get sheets
        sheet_players = wb["Giocatori"]
        sheet_matches = wb["Partite"]
        sheet_summary = wb["Reassunto"]
        
        # Create mappings of player names to row indices in the Players sheet
        player_to_row = {}
        for row in range(2, len(all_players) + 2):  # +2 because Excel is 1-indexed and has a header row
            player_name = sheet_players.cell(row=row, column=1).value
            player_to_row[player_name] = row
        
        # Add cell references in Matches sheet
        for row in range(2, len(self.matches) + 2):
            for col in range(4, 8):  # Columns D, E, F, G (player columns)
                player_name = sheet_matches.cell(row=row, column=col).value
                if player_name in player_to_row:
                    player_row = player_to_row[player_name]
                    # Set formula reference
                    sheet_matches.cell(row=row, column=col).value = f'=SE(Giocatori!A{player_row}="","",Giocatori!A{player_row})'
        
        # Create formulas for summary sheet
        for row in range(2, len(all_players) + 2):
            player_name = sheet_summary.cell(row=row, column=1).value
            if player_name in player_to_row:
                player_row = player_to_row[player_name]
                # Set formula reference for name
                sheet_summary.cell(row=row, column=1).value = f'=SE(Giocatori!A{player_row}="","",Giocatori!A{player_row})'
                
                # Correzione alla formula per il calcolo dei games vinti
                # Quando sono nella squadra 1 (giocatore 1 o 2): games vinti - games persi
                # Quando sono nella squadra 2 (giocatore 3 o 4): games vinti - games persi
                
                formula_games_vinti = (
                    # Quando sono giocatore 1 o 2 (Squadra 1)
                    f'=SE(Giocatori!A{player_row}="","",SUMIFS(Partite!H:H,Partite!D:D,Giocatori!A{player_row})+' +  # Games vinti come giocatore 1
                    f'SUMIFS(Partite!H:H,Partite!E:E,Giocatori!A{player_row})+' +  # Games vinti come giocatore 2
                    
                    # Quando sono giocatore 3 o 4 (Squadra 2)
                    f'SUMIFS(Partite!I:I,Partite!F:F,Giocatori!A{player_row})+' +  # Games vinti come giocatore 3
                    f'SUMIFS(Partite!I:I,Partite!G:G,Giocatori!A{player_row}))'  # Games vinti come giocatore 4
                )

                formula_partite_vinti = (
                    f'=SE(Giocatori!A{player_row}="","",SUMPRODUCT(('
                    f'(Partite!D:D=Giocatori!A{player_row})+(Partite!E:E=Giocatori!A{player_row}))'
                    f'*(Partite!H:H>Partite!I:I)) + '
                    f'SUMPRODUCT(('
                    f'(Partite!F:F=Giocatori!A{player_row})+(Partite!G:G=Giocatori!A{player_row}))'
                    f'*(Partite!I:I>Partite!H:H)))'
                )

                formula_games_persi = (
                    # Quando sono giocatore 1 o 2 (Squadra 1)
                    f'=SE(Giocatori!A{player_row}="","",SUMIFS(Partite!I:I,Partite!D:D,Giocatori!A{player_row})+' +  # Games persi come giocatore 1
                    f'SUMIFS(Partite!I:I,Partite!E:E,Giocatori!A{player_row})+' +  # Games persi come giocatore 2
                    
                    # Quando sono giocatore 3 o 4 (Squadra 2)
                    f'SUMIFS(Partite!H:H,Partite!F:F,Giocatori!A{player_row})+' +  # Games persi come giocatore 3
                    f'SUMIFS(Partite!H:H,Partite!G:G,Giocatori!A{player_row}))'     # Games persi come giocatore 4
                )

                formula_differenza = (
                    # Quando sono giocatore 1 o 2 (Squadra 1)
                    f'=SE(Giocatori!A{player_row}="","",SUMIFS(Partite!H:H,Partite!D:D,Giocatori!A{player_row})+' +  # Games vinti come giocatore 1
                    f'SUMIFS(Partite!H:H,Partite!E:E,Giocatori!A{player_row})-' +  # Games vinti come giocatore 2
                    f'SUMIFS(Partite!I:I,Partite!D:D,Giocatori!A{player_row})-' +  # Games persi come giocatore 1
                    f'SUMIFS(Partite!I:I,Partite!E:E,Giocatori!A{player_row})+' +  # Games persi come giocatore 2
                    
                    # Quando sono giocatore 3 o 4 (Squadra 2)
                    f'SUMIFS(Partite!I:I,Partite!F:F,Giocatori!A{player_row})+' +  # Games vinti come giocatore 3
                    f'SUMIFS(Partite!I:I,Partite!G:G,Giocatori!A{player_row})-' +  # Games vinti come giocatore 4
                    f'SUMIFS(Partite!H:H,Partite!F:F,Giocatori!A{player_row})-' +  # Games persi come giocatore 3
                    f'SUMIFS(Partite!H:H,Partite!G:G,Giocatori!A{player_row}))'      # Games persi come giocatore 4
                )
                sheet_summary.cell(row=row, column=2).value = formula_games_vinti # Games vinti
                sheet_summary.cell(row=row, column=3).value = formula_partite_vinti# Partite vinti
                sheet_summary.cell(row=row, column=4).value = formula_games_persi # Games persi
                sheet_summary.cell(row=row, column=5).value = formula_differenza # Differenza
                
        
        # Set protection with password [Disabled for now]
        #for sheet in wb.sheetnames:
        #    wb[sheet].protection.sheet = True
        #    wb[sheet].protection.password = "P4D3L1ND**RPADIGLIONE00042"
        
        # Save workbook
        wb.save(self.export_path)
        print(f"Tournament Excel file generated successfully at {self.export_path}")

def parse_time(time_str):
    """Parse time string into datetime object."""
    return datetime.strptime(time_str, "%H:%M")

def main():
    # Example usage with command-line arguments
    if len(sys.argv) < 7:
        print("Usage: python kingone-generator.py <duration_minutes> <match_duration> <start_time> <num_courts> <num_players> <export_path>")
        print("Example: python kingone-generator.py 120 20 18:00 4 16 ./torneo_padel.xlsx")
        return
    
    try:
        duration_minutes = int(sys.argv[1])
        match_duration = int(sys.argv[2])
        start_time = parse_time(sys.argv[3])
        num_courts = int(sys.argv[4])
        num_players = int(sys.argv[5])
        export_path = sys.argv[6]
        
        # Create generator and export Excel
        generator = PadelTournamentGenerator(
            duration_minutes=duration_minutes,
            match_duration=match_duration,
            start_time=start_time,
            num_courts=num_courts,
            num_players=num_players,
            export_path=export_path
        )
        
        generator.generate_excel()
        
    except ValueError as e:
        print(f"Errore: {str(e)}")
    except Exception as e:
        print(f"Si è verificato un errore: {str(e)}")

if __name__ == "__main__":
    main()